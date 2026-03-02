"""Excel 解析器"""

from openpyxl import load_workbook


def parse_excel(file_path: str) -> str:
    """
    解析 Excel 文件
    
    Args:
        file_path: .xlsx 文件路径
        
    Returns:
        提取的文本内容
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    texts = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        texts.append(f"=== 工作表: {sheet_name} ===")

        rows_data = []
        for row in sheet.iter_rows(values_only=True):
            # 过滤空行
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(v.strip() for v in row_values):
                rows_data.append(row_values)

        if rows_data:
            # 第一行作为表头
            header = rows_data[0] if rows_data else []
            if header:
                texts.append("表头: " + " | ".join(header))
                texts.append("-" * 40)

            for i, row in enumerate(rows_data[1:], 1):
                if header:
                    row_text = " | ".join(
                        f"{h}: {v}" for h, v in zip(header, row) if v.strip()
                    )
                else:
                    row_text = " | ".join(row)
                if row_text:
                    texts.append(f"行 {i}: {row_text}")

        texts.append("")

    wb.close()
    return "\n".join(texts)
